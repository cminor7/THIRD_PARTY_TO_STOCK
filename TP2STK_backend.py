# ver 2.0.1
# internal libraries
from os import path, getlogin, makedirs
from datetime import datetime
from pathlib import Path
from shutil import rmtree, move
from re import sub
from getpass import getuser

# external libraries
import pandas as pd
import numpy as np
import openpyxl as xl
import win32com.client as win32
import pyodbc
import snowflake.connector as sf
from SMTP_email import *


def getContact():

	# server = 'localhost\sqlexpress' # for a named instance
	# server = 'myserver,port' # to specify an alternate port
	server, database, username, password = 'server', 'database', '', '' 
	cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+password)
	cursor = cnxn.cursor()
	query = open('DEVELOPER_FILES/contact_query.sql', 'r').read()
	return pd.read_sql(query, cnxn)


def sf_con():
    conn = sf.connect(
    user=getuser(),
    account='',
    password='',
    authenticator="",
    token='',
    warehouse='',
    database='',
    schema='',
    role=''
    )
    return conn


# global variables 
EMAIL_MESSAGE = 'DEVELOPER_FILES/TP2STK Supplier E-mail.txt'
CURRENT_DATE = datetime.today().strftime('%Y-%m-%d')
DF_CONTACT = getContact()


def userEmail():

	# outlook must be re-opened for this to work, use to get user's email address
	# remove gen_py folder
	user_id = getlogin()
	gen_py = f'C:/Users/{user_id}/AppData/Local/Temp/gen_py'
	if path.exists(gen_py):
   		rmtree(gen_py)
	outlook = win32.gencache.EnsureDispatch("Outlook.Application")
	return outlook.Session.CurrentUser.AddressEntry.GetExchangeUser().PrimarySmtpAddress


def splitFiles():

	print('\nCREATING TEMPLATE FILES FOR EACH SUPPLIER...\n')
	
	sql_split = """SELECT DISTINCT MATERIAL, SUPPLIER_NO, SUPPLIER_STOCK_NUMBER, MATERIAL_DESCRIPTION, SUPPLIER_NAME, ORDER_QUANTITY, CURRENT_SLT, CURRENT_IBQ, DROP_SHIP_100
		
		FROM ISP.RA.TP2STK 
		WHERE (DATE_SENT_TO_SUPPLIER IS NULL AND DCET_APPROVAL = 'Y') 
		--OR (DATE_SENT_TO_SUPPLIER >= '2022-01-01' AND CURRENT_DATE() - DATE_SENT_TO_SUPPLIER > 30 AND SUPPLIER_APPROVAL IS NULL AND DCET_APPROVAL = 'Y')
	"""
	df = pd.read_sql(sql_split, sf_con(), index_col=None) # main dataframe - filter out rows we don't need
	supplier_list = df['SUPPLIER_NO'].unique().tolist()

	file_created = 0
	missing_contact = []
	for supplier_no in supplier_list:

		if supplier_no not in DF_CONTACT['SUPPLIER'].values:
			missing_contact.append(supplier_no)
			continue

		wb = xl.load_workbook('DEVELOPER_FILES/TP2STK Supplier Feedback Template.xlsx', read_only=False)
		ws = wb['ELIGIBLE ITEMS']

		current_row = 2
		current_supplier = df[df['SUPPLIER_NO'] == supplier_no]
		current_supplier = current_supplier.reset_index()
		current_supplier_contact = DF_CONTACT[DF_CONTACT['SUPPLIER'] == supplier_no]

		# writing to template
		for row, col in current_supplier.iterrows():

			ws[f'A{current_row}'] = col['MATERIAL']
			ws[f'B{current_row}'] = col['SUPPLIER_STOCK_NUMBER']
			ws[f'C{current_row}'] = col['MATERIAL_DESCRIPTION']
			ws[f'D{current_row}'] = col['SUPPLIER_NAME']
			ws[f'E{current_row}'] = col['ORDER_QUANTITY']
			ws[f'F{current_row}'] = col['CURRENT_SLT']
			ws[f'G{current_row}'] = col['CURRENT_IBQ']
			ws[f'H{current_row}'] = col['DROP_SHIP_100']
			ws[f'T{current_row}'] = col['SUPPLIER_NO']
			ws[f'U{current_row}'] = current_supplier_contact.iloc[0]['SUPPLIER_PERFORMANCE_ANALYST']
			current_row += 1

		filename = 'SPLIT_FILES/' + str(int(current_supplier.at[0, 'SUPPLIER_NO'])) + '_GRAINGER_THIRD PARTY TO STOCK.xlsx'
		wb.save(filename)
		wb.close()
		file_created += 1
	print(f'\n{file_created} TEMPLATE FILES CREATED.\n')
	print(f'\nMISSING CONTACT LIST: {missing_contact}\n')


def sendSMTP():

	print('\nSENDING SUPPLIER SMTP EMAILS...\n')

	cur = sf_con().cursor()

	# getting a list of file path location
	files = Path('SPLIT_FILES').glob('*.xlsx')
	file_list = [path.abspath(filepath) for filepath in files]

	# create a new folder to move current split files after sending email
	new_folder = f'SPLIT_FILES/ARCHIVE/{CURRENT_DATE}'
	if not path.exists(new_folder):
   		makedirs(new_folder)
	 
	message = open(EMAIL_MESSAGE,'r', encoding='utf-8').read()
	server_success_response = open('DEVELOPER_FILES/server_success_response.txt', 'r').read()
	user_email = userEmail()
	email_sent = 0
	email_error = 0

	for file in file_list:

		df_split_file = pd.read_excel(file, sheet_name='ELIGIBLE ITEMS', engine='openpyxl')
		last_row = df_split_file['MATERIAL'].count() + 1

		wb = xl.load_workbook(file, read_only=False)
		ws = wb['ELIGIBLE ITEMS']

		# write the date_sent_to_supplier column in split files
		for cell in ws[f'V2:V{last_row}']:
			cell[0].value = datetime.today().strftime('%m/%d/%Y')

		wb.save(file)
		wb.close()

		# lookup the contact info for current template file
		current_supplier_no = int(Path(file).stem.split('_')[0])

		if not current_supplier_no in DF_CONTACT['SUPPLIER'].values:
			email_error += 1
			continue

		filter_contact = DF_CONTACT[DF_CONTACT['SUPPLIER'] == current_supplier_no]
		supplier_email_list = list(filter(None,filter_contact['SUPPLIER_EMAIL'].unique().tolist()))
		supplier_email = ';'.join(supplier_email_list)
		spa_email = filter_contact.iloc[0]['SPA_EMAIL']
		supplier_name = filter_contact.iloc[0]['SUPPLIERNAME'].strip().replace('.', '')
		supplier_number = filter_contact.iloc[0]['SUPPLIER']

		server_response = SMTP(em_from=spa_email, 
			em_to=supplier_email, 
			em_cc=f'{user_email};{spa_email}'.lower().replace('supplier_followup@grainger.com', ''),
			em_subject=f'ACTION REQUIRED: THIRD PARTY TO STOCK - {supplier_name}', 
			em_message=message, 
			em_attachment=file)

		try:
			server_error = serverError(response=server_response, success_response=server_success_response)
		except:
			server_error = 'OTHER ERROR'

		if server_error == 'NO ERROR':
			# set date_sent_to_supplier values in database for successful emails
			material_sent = df_split_file.loc[df_split_file['SUPPLIER NUMBER'] == current_supplier_no, 'MATERIAL'].tolist()
			material_list = ', '.join(f"'{material}'" for material in material_sent)
			sql_date_sent = f"UPDATE ISP.RA.TP2STK SET DATE_SENT_TO_SUPPLIER = '{CURRENT_DATE}' WHERE MATERIAL IN ({material_list})"
			cur.execute(sql_date_sent)

			moveFile(file=file, dst_folder=new_folder)
			email_sent += 1
		else:
			email_error += 1			

	print(f'\n{email_sent} EMAILS SENT, {email_error} EMAILS FAILED.\n')


def sendOutlook():

	print('\nSENDING SUPPLIER OUTLOOK EMAILS...\n')

	cur = sf_con().cursor()
	# getting a list of file path location
	files = Path('SPLIT_FILES').glob('*.xlsx')
	file_list = [path.abspath(filepath) for filepath in files]

	# create a new folder to move current split files after sending email
	new_folder = f'SPLIT_FILES/ARCHIVE/{CURRENT_DATE}'
	if not path.exists(new_folder):
   		makedirs(new_folder)
	 
	message = open(EMAIL_MESSAGE,'r', encoding='utf-8').read()
	user_email = userEmail()
	email_sent = 0
	email_error = 0

	for file in file_list:

		df_split_file = pd.read_excel(file, sheet_name='ELIGIBLE ITEMS', engine='openpyxl')
		last_row = df_split_file['MATERIAL'].count() + 1

		wb = xl.load_workbook(file, read_only=False)
		ws = wb['ELIGIBLE ITEMS']

		for cell in ws[f'V2:V{last_row}']:
			cell[0].value = datetime.today().strftime('%m/%d/%Y')

		wb.save(file)
		wb.close()

		# lookup the contact info for current template file
		current_supplier_no = int(Path(file).stem.split('_')[0])

		if not current_supplier_no in DF_CONTACT['SUPPLIER'].values:
			email_error += 1
			continue

		filter_contact = DF_CONTACT[DF_CONTACT['SUPPLIER'] == current_supplier_no]
		supplier_email_list = list(filter(None,filter_contact['SUPPLIER_EMAIL'].unique().tolist()))
		supplier_email = ';'.join(supplier_email_list)
		spa_email = filter_contact.iloc[0]['SPA_EMAIL']
		supplier_name = filter_contact.iloc[0]['SUPPLIERNAME'].strip().replace('.', '')

		outlook(em_to=supplier_email, 
			em_cc=f'{user_email};{spa_email}'.lower().replace('supplier_followup@grainger.com', ''), 
			em_subject=f'ACTION REQUIRED: THIRD PARTY TO STOCK - {supplier_name}', 
			em_message=message, 
			em_attachment=file)

		material_sent = df_split_file.loc[df_split_file['SUPPLIER NUMBER'] == current_supplier_no, 'MATERIAL'].tolist()
		material_list = ', '.join(f"'{material}'" for material in material_sent)
		sql_date_sent = f"UPDATE ISP.RA.TP2STK SET DATE_SENT_TO_SUPPLIER = '{CURRENT_DATE}' WHERE MATERIAL IN ({material_list})"
		cur.execute(sql_date_sent)

		moveFile(file=file, dst_folder=new_folder)
		email_sent +=1

	print(f'\n{email_sent} EMAILS SENT, {email_error} EMAILS FAILED.\n')


def stitchFiles():

	print('\nIMPORTING SUPPLIER FEEDBACK...\n')

	cur = sf_con().cursor()
	files = Path('IMPORT_FILES').glob('*.xlsx')
	file_list = [path.abspath(filepath) for filepath in files]
	
	# get all the import files into one dataframe
	if len(file_list) > 1:
		import_files = pd.concat((pd.read_excel(f, sheet_name='ELIGIBLE ITEMS', engine='openpyxl') for f in file_list), ignore_index=True).dropna(how='all').drop_duplicates()
	else:
		import_files = pd.read_excel(file_list[0], sheet_name='ELIGIBLE ITEMS', engine='openpyxl').dropna(how='all').drop_duplicates()

	
	import_files['COMMENTS'] = [sub(r'[^A-Za-z0-9?!,.:; ]+','', str(x)) for x in import_files['COMMENTS']] # remove new character line to prevent import errors
	import_files['DO NOT TRANSITION REASON*'] = [sub(r'[^A-Za-z0-9?!,.:; ]+','', str(x)) for x in import_files['DO NOT TRANSITION REASON*']] # remove new character line to prevent import errors
	import_files[['SUPPLIER CONFIRMED ORDER MULTIPLE*', 'SUPPLIER CONFIRMED LEAD TIME*']] = import_files[['SUPPLIER CONFIRMED ORDER MULTIPLE*', 'SUPPLIER CONFIRMED LEAD TIME*']].apply(pd.to_numeric, errors='coerce')
	import_files['MATERIAL'] = import_files['MATERIAL'].astype(str)

	import_files_list_tuple = list(zip(*map(import_files.get, ['MATERIAL', 'TRANSITION STATUS*', 'SUPPLIER CONFIRMED ORDER MULTIPLE*', 
		'SUPPLIER CONFIRMED LEAD TIME*', 'COMMENTS', 'DO NOT TRANSITION REASON*']))) # convert the rows of dataframe with the selected column into a list of tuples
	import_files_tuple_str = ', '.join(map(str, import_files_list_tuple)).replace("'nan'", 'NULL').replace("nan", 'NULL') # converts the list of tuples into one string for SQL update command below

	update_query = f"""UPDATE ISP.RA.TP2STK t
    SET SUPPLIER_APPROVAL = (CASE WHEN v.supplier_approval = 'Transition' THEN 'Y' ELSE 'N' END),
    	NEW_IBQ = (CASE WHEN v.supplier_approval = 'Transition' AND v.new_ibq IS NOT NULL THEN v.new_ibq ELSE NULL END),
        NEW_SLT = (CASE WHEN v.supplier_approval = 'Transition' AND v.new_slt IS NOT NULL THEN v.new_slt ELSE NULL END),
        GENERAL_REJECTION_REASON = (CASE WHEN v.supplier_approval = 'DoNotTransition' THEN LEFT(TRIM(v.general_rejection_reason), 500) ELSE NULL END),
        COMMENTS = LEFT(TRIM(v.comments), 1200),
        ORDER_QUANTITY = (CASE WHEN v.new_ibq IS NOT NULL AND v.supplier_approval = 'Transition' THEN 
        						GREATEST(ROUND(DIV0NULL(t.R12_DEMAND_UNITS, 12) * 2 / v.new_ibq, 0) * v.new_ibq, v.new_ibq)
        					ELSE 
        						GREATEST(ROUND(DIV0NULL(t.R12_DEMAND_UNITS, 12) * 2 / t.CURRENT_IBQ, 0) * t.CURRENT_IBQ, t.CURRENT_IBQ)
        				END)
    FROM 
    (
        VALUES {import_files_tuple_str}
    ) v (material, supplier_approval, new_ibq, new_slt, comments, general_rejection_reason)
    
    WHERE t.MATERIAL = v.material
	"""
	cur.execute(update_query)

	# create a new folder to move current import files after importing
	new_folder = f'IMPORT_FILES/ARCHIVE/{CURRENT_DATE}'
	if not path.exists(new_folder):
   		makedirs(new_folder)

	for file in file_list:
		try:
			moveFile(file=file, dst_folder=new_folder)
		except:
			continue

	print(f'\nIMPORT COMPLETE.')


def moveFile(file, dst_folder):

	# rename the file before moving if filename already taken in folder
	file_stem = sub(r'[\(\[].*?[\)\]]', '', Path(file).stem).strip()
	new_file_path = f'{dst_folder}/{file_stem}.xlsx'
	file_name_exist = path.isfile(new_file_path)
	counter = 1

	while file_name_exist:
		new_file_path = f'{dst_folder}/{file_stem}({counter}).xlsx'
		file_name_exist = path.isfile(new_file_path)
		counter += 1

	move(file, new_file_path)


def serverError(response, success_response):
	if success_response == response:
		error_response = 'NO ERROR'
	else:
		error_response = 'OTHER SERVER ERROR'
	return error_response